import re
from openpyxl import load_workbook
from itertools import chain
import spacy

nlp = spacy.load("en_core_web_sm")
# extract link from the text in csv file
# def extract_link(text):
#     # regex to extract link from text
#     regex = r"https?://[^\s]+"
#     # findall() has been used
#     # with valid conditions for urls in string
#     url = re.findall(regex, text)
#     return url

def extract_link():
    # load the excel file
    wb = load_workbook("C:\\Users\\Kshitiz Pandya\\Downloads\\tweets.xlsx")
    ws = wb.active
    institution_name = ws['D']
    url = []
    for x in range(len(institution_name)):
        regex = r"https?://[^\s]+"
        a = re.findall(regex, institution_name[x].value)
        url.append(a)
    # print(url)

    res = [ele for ele in url if ele != []]

    # printing result
    # print(res)

    flatten_list = list(chain.from_iterable(res))
    print(len(flatten_list))
    print(flatten_list)

def cleaning():
    # clean a string from excel file and for special characters
    # load the excel file
    wb = load_workbook("C:\\Users\\Kshitiz Pandya\\Downloads\\tweets.xlsx")
    ws = wb.active
    institution_name = ws['D']
    clean = []
    for i in range(len(institution_name)):

        new_text = re.sub(r"https?://[^\s]+", '', institution_name[i].value)
        new_text1 = re.sub(r'[^\w\s]', '', new_text)
        clean.append(new_text1)
        # print(i, ".", new_text)
    # print(len(clean))
    clean = clean[1:19]
    # print(len(clean))
    print(clean)

    final = []
    for i in range(len(clean)):
        doc = nlp(clean[i])
        tokens = [token.text for token in doc]
        # print(tokens)
        q = tokens[0]
        w = tokens[1]
        e = tokens[2]
        r = tokens[3]
        t = tokens[4]
        y = tokens[5]
        ans = q + " " + w + " " + e + " " + r + " " + t + " " + y
        final.append(ans)
    print(final)

# r'[^\w\s]'
extract_link()
cleaning()