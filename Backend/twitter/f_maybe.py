import re
from openpyxl import load_workbook
from itertools import chain

# extract link from the text in csv file
# def extract_link():
#     # load the excel file
#     wb = load_workbook("C:\\Users\\Kshitiz Pandya\\Downloads\\passionate coder.xlsx")
#     ws = wb.active
#     text = ws['E']
#     url = []
#     for x in range(len(text)):
#         regex = r"https?://[^\s]+"
#         a = re.findall(regex, text[x].value)
#         url.append(a)
#     print(url)
#
#     res = [ele for ele in url if ele != []]
#
#     # printing result
#     # print(res)
#
#     flatten_list = list(chain.from_iterable(res))
#     print(flatten_list)

def cleaning():
    # clean a string from excel file and for special characters
    # load the excel file
    wb = load_workbook("C:\\Users\\Kshitiz Pandya\\Downloads\\passionate coder.xlsx")
    ws = wb.active
    text = ws['E']
    demo = []

    for i in range(len(text)):
        z = text[i].value
        z = z.strip('\n')
        new_text = re.sub(r"https?://[^\s]+", '', z)
        new_text = re.sub(r"#", '', new_text)
        new_text = re.sub(r"@", '', new_text)
        # new_text = re.sub(r"\[", '', new_text)
        # new_text1 = re.sub(r'[^\w\s]', '', new_text)
        # new_text.strip('\n')
        print(i, ".", new_text)
        demo.append(new_text)
    print(demo)

# extract_link()
cleaning()